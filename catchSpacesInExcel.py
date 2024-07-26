import openpyxl as xl
import string

iterative_dict, numero, main_dict_w_answers, dict_index, comparison_sheets = {}, 0, {}, {},{}

donor = input("Please give me a filepath")
xlobject = xl.load_workbook(donor)

recipient = input("Please give me the filepath to the document that you would like to recieve the answers")
comparison = xl.load_workbook(recipient)

for i in xlobject.sheetnames:
	numero += 1
	iterative_dict[numero] = i

print("Sheet names to index from")
[print(i, "-", iterative_dict[i]) for i in iterative_dict]
index = int(input("Please enter an number to index the appropriate sheet"))

grimbsy = xlobject[iterative_dict[index]]	#getting the dictionary

def get_merged_cells_values(sheet,cell):
	for merged_cell in sheet.merged_cells.ranges:
		if cell.coordinate in merged_cell:
			value = sheet.cell(merged_cell.min_row, merged_cell.min_col).value
			print("The cell {cell.coordinate} is merged. The value is {cell.value}")	#just a bit of a sanity check here
			return value
	return cell.value

for i in range(1, grimbsy.max_row + 1):
	for j in range(grimbsy.max_column + 1):
		if str(grimbsy.cell(row=i, column=j).coordinate)[0]	=="B" and grimbsy.cell(row=i, column=j).value is not None and i != 1:
			main_dict_w_answers[f"{str(grimbsy.cell(row=i,column=j).value) + str(grimbsy.cell(row=i, column=j)).coordinate}"] = [str(grimbsy.cell(row=i,column=j + 1).value),str(grimbsy.cell(row=i,column=j + 2).value), str(grimbsy.cell(row=i,column=j + 3).value)]
		if str(grimbsy.cell(row=i, column=j).coordinate)[0]	== "A" and grimbsy.cell(row=i, column=j + 1).value is not None and i != 1 and grimbsy.cell(row=i, column=j).value is not None and i != 1:
			main_dict_w_answers[f"{str(grimbsy.cell(row=i, column=j).value) + str(grimbsy.cell(row=i, column=j).coordinate)}"] = [str(grimbsy.cell(row=i,column=j + 2).value),str(grimbsy.cell(row=i,column=j + 3).value), str(grimbsy.cell(row=i,column=j + 4).value)]

persons_index = list(set(i[:-3] for i in main_dict_w_answers))

numero = 0

for i in persons_index:
	numero += 1
	dict_index[numero] = i

[print(i, "-", dict_index[i] for i in dict_index)]
pick = int(input("Please select a number to index the applications please"))

[print(i, "-", main_dict_w_answers[i][0], main_dict_w_answers[i][1], main_dict_w_answers[i][2]) for i in main_dict_w_answers if dict_index[pick] in i]	#sanity check

remove_punctuation = lambda s:s.lower().translate(str.maketrans('','', string.punctiation))

for i in comparison.sheetnames:
	numero += 1
	comparison_sheets[numero]

sheet_to_compare_against = comparison[comparison_sheets[index]]

for n in range(1, sheet_to_compare_against.max_row +1):
	for p in range(1, sheet_to_compare_against.max_column + 1):
		if "?" in str(sheet_to_compare_against.cell(row=n,column=p)):
			for i in main_dict_w_answers:
				if dict_index[pick] in i:
					if str(remove_punctuation(sheet_to_compare_against.cell(row=n,column=p).value)) in remove_punctuation(main_dict_w_answers[i][1]):
						if str(remove_punctuation(sheet_to_compare_against.cell(row=n,column=p+1).value)) is None:
							sheet_to_compare_against[f"{sheet_to_compare_against.cell(row=n,column=p+1).coordinate}"] = str(main_dict_w_answers[i][2])
							print(f"Writing {str(main_dict_w_answers[i][2])} to {sheet_to_compare_against[{sheet_to_compare_against.cell(row=n,column=p+1).coordinate}]}")
							comparison.save(recipient)
							completion = 1

if completion != 1:
	for n in range(1, sheet_to_compare_against.max_row +1):
		for p in range(1, sheet_to_compare_against.max_column + 1):
			if str(sheet_to_compare_against.cell(row=n, column=p).coordinate)[0] == "B" and n != 1:
				for i in main_dict_w_answers:
					if str(dict_index[pick]) in str(i):
						if remove_punctuation(sheet_to_compare_against.cell(row=n, column=p + 1).value) in remove_punctuation(main_dict_w_answers[i][1]):
							if sheet_to_compare_against.cell(row=n, column=p+1).value is None:
								sheet_to_compare_against[f"{sheet_to_compare_against.cell(row=n,column=p+1).coordinate}"] = str(main_dict_w_answers[i][2])
								comparison.save(recipient)
								print(f"Answer was writtten to cell {sheet_to_compare_against.cell(row=n,column=p+1).coordinate}")
							else:
								print(f"Answer was not writtten to cell {sheet_to_compare_against.cell(row=n,column=p+1).coordinate} as data exists in that cell")