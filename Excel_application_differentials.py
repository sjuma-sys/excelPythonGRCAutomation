import openpyxl as xl
import string 

iterative_dict,numero,main_dict_w_answers,dict_index,comparison_sheets = {},0,{},{},{}


donor = input("Please give me the filepath for the donor")
xlobject = xl.load_workbook(donor)

recipient = input("Please give me the filepath for the recipient that you would like to recieve the answers")
comparison = xl.load_workboo(recipient)

for i in xlobject.sheetnames:
	numero +=1
	iterative_dict[numero] = i

numero = 0

print("Please index from a sheet name displayed here")
for i in iterative_dict:
	print(i, iterative_dict[i])

index = int(input("Please enter a number to select a sheet"))

grimbsy = xlobject[iterative_dict[index]]	

for i in range(1, grimbsy.max_row + 1):
	for j in range(1, grimbsy.max_column + 1):
		if str(grimbsy.cell(row=i,column=j).value) == "B" and grimbsy.cell(row=i,column=j) is not None and i != 1:
			main_dict_w_answers[f"{str(grimbsy.cell(row=i,column=j).value) + str(grimbsy.cell(row=i,column=j).coordinate)}"] = [str(grimbsy.cell(row=i,column=j+1).value), str(grimbsy.cell(row=i,column=j+2).value), str(grimbsy.cell(row=i,column=j+3).value)]


persons_index = list(set(i[:-3] for i in main_dict_w_answers))

numero = 0

for i in persons_index:
	numero +=1
	dict_index[numero]  = i

for i in dict_index:
	print(i, "-", dict_index[i])

pick = int(input("Please select a number to index the applications please"))

numero = 0

for i in comparison.sheetnames:
	numero +=1
	comparison_sheets[numero] = i


for i in comparison_sheets:
	print(i, comparison_sheets[i])

index - int(input("Please tell me which worksheet you are inputting the informtion into"))

sheet_to_compare_against = comparison[comparison_sheets[index]]

remove_punctuation = lambda s: s.lower().translate(str.maketrans('','', string.punctuation))

for n in range(1, sheet_to_compare_against.max_row + 1):
	for p in range(1, sheet_to_compare_against.max_column + 1):
		if "?" in str(sheet_to_compare_against.cell(row=n, column=p).values):
			for i in main_dict_w_answers:
				if dict_index[pick] in i:
					if str(remove_punctuation(sheet_to_compare_against.cell(row=n,column=p).value)) in str(remove_punctuation(main_dict_w_answers[i][1])):
						print(f"Answers from comparison file: {sheet_to_compare_against.cell(row=n, column=p).value}, Answers from the original: {main_dict_w_answers[i][2]} \nCell to be written to: {sheet_to_compare_against.cell(row=n,column=p + 1).coordinate}")
						if sheet_to_compare_against.cell(row=n,column=p+1).value is None:
							sheet_to_compare_against[f"{sheet_to_compare_against.cell(row=n, column=p+1).coordinate}"] = str{main_dict_w_answers[i][2]}
							comparison.save(recipient)


