import openpyxl as xl
import string

numero,iterative_dict,values_and_coords_file_1,dict_to_load_answers_into,comparison_sheets  = 0,{},{},{},{}

filepath_of_donor = input("Give me a path to the donor file (The file that contains the answers)")

filepath_of_recipient = input("Give me a path to the recipient file (The file that contains the unanswered questions)")

print("Make sure the file you are saving to is closed or else the script will not work")

donor = input("Please give me the filepath for the donor")
xlobject = xl.load_workbook(donor)

recipient = input("Please give me the filepath for the recipient that you would like to recieve the answers")
comparison = xl.load_workboo(recipient)


for i in xlobject.sheetnames:
	numero +=1
	iterative_dict[numero] = i

numero = 0

[print(f"{i} \t {iterative_dict[i]}") for i in iterative_dict]

index = int(input("Please give me an int to index the sheetname from the donor file"))

grimbsy = xlobject[iterative_dict[index]]

for i in range(1, grimbsy.max_row + 1):
	for j in range(1, grimbsy.max_column):
		if "?" in str(grimbsy.cell(row=i,column=j).value):
			values_and_coords_file_1[grimbsy.cell(row=i,column=j).value] = grimbsy.cell(row=i, column=j).value

numero = 0

for i in comparison.sheetnames:
	numero += 1
	comparison[comparison_sheets[index]]

[print(f"{i} \t {iterative_dict[i]}") for i in comparison_sheets]

index = input("Select a number to index a relevant sheet")

sheet_to_compare_against = comparison[comparison_sheets[index]]

remove_punctuation = lambda s: s.lower().translate(str.maketrans('','', string.punctuation))

for n in range(1, sheet_to_compare_against.max_row + 1):
	for p in range(1, sheet_to_compare_against.max_column + 1):
		if "?" in str(sheet_to_compare_against.cell(row=n, column=p).value):
			dict_to_load_answers_into[sheet_to_compare_against.cell(row=n, column=p).value] = sheet_to_compare_against.cell(row=n, column=p).coordinate
			for i in values_and_coords_file_1:
				if str(remove_punctuation(i)) in str(remove_punctuation(dict_to_load_answers_into.cell(row=n, column=p).value)) or str(remove_punctuation(values_and_coords_file_1[i])) in str(remove_punctuation(dict_to_load_answers_into.cell(row=n, column=p).value))	#mabye incorrect and set I to index into dict or to 
					print(f"Question is {dict_to_load_answers_into.cell(row=n, column=p).value} Answer is {values_and_coords_file_1[i]}")


for key1 in dict_to_load_answers_into:
	for key2 in values_and_coords_file_1:
		if remove_punctuation(key1) in remove_punctuation(key2):
			if sheet_to_compare_against[dict_to_load_answers_into[key1]].value is None:
				print(f"Going into: {sheet_to_compare_against[dict_to_load_answers_into[key1]]} is the value {values_and_coords_file_1[key2]}")
				#sheet_to_compare_against[dict_to_load_answers_into[key1]] =  values_and_coords_file_1[key2]

